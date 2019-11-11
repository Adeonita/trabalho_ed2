import openpyxl, os  

arquivo = openpyxl.load_workbook('dados.xlsx') 
planilhas = arquivo.get_sheet_names()
planilha = arquivo.get_sheet_by_name(planilhas[0])
sheet = arquivo.active
numero_de_linhas = planilha.max_row
numero_de_colunas = planilha.max_column
sheet = arquivo.active

indice = 2 #Começa em 2 pois a primeira linha está vazia
campos = sheet['A2'].value #Atribui a linha 2 (campos) a variavel campos
campos = campos.split(';')

#Retorna todas as linhas do arquivo
for variavel in range(indice, numero_de_linhas):
    a1 = sheet['A{}'.format(numero_de_linhas)]
    print('{}'.format(a1.value))
print(campos[0])

#a1 = sheet['A2'].value
#print(a1)
#a1 = sheet['A2']
#a2 = sheet['A3']
#print(a1.value)   
#print(a2.value)

